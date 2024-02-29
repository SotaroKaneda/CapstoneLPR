import openpyxl


workbook = openpyxl.load_workbook(r"D:\v2x-11-30-data\ALPRPlateExport11-30-23Revised.xlsx")
sheet = workbook.active
sheet1 = workbook["Sheet1"]
jurisdiction_dict = {}

# parse sheet one for jurisdictions
for row in sheet1.iter_rows(min_row=2, max_row=197137, values_only=True):
    UFM_ID = row[0]
    PLATE_JURISDICTION = row[6]
    jurisdiction_dict[UFM_ID] = PLATE_JURISDICTION

with open(r"C:\Users\Jed\Desktop\capstone_project\v2x-dataset\data-11-30.csv", "w") as file:
    headers = ",".join(list(next(sheet.values))) + ",\n"
    file.write(headers)
    for row in sheet.iter_rows(min_row=2, max_row=55345, values_only=True):
        row = list(row)
        row[8] = jurisdiction_dict[row[0]]
        row[11] = row[11].split("/")[-1]
        if row[12] : row[12] = row[12].split("/")[-1]
        if row[13] : row[13] = row[13].split("/")[-1]
        if row[14] : row[14] = row[14].split("/")[-1]

        # update jurisiction here. The datasheet has the excel formula instead of the jurisdiction
        write_line = ""
        for entry in row:
            write_line += f"{str(entry)},"

        write_line += "\n"
        file.write(write_line)

